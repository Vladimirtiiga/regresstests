from openpyxl import load_workbook
from startgen import test_get_request
from getisightstatus import get_status
from reportcompare import*
import time
import os
import json  # Для парсинга вложенных фрагментов, если нужно
from cromeget2 import getupdate
from datetime import datetime, date, time as dt
getupdate()
server='dev'
#server='prod'
if server=='dev':
    token = os.getenv('tokendev')
if server == 'prod':
    token = os.getenv('tokenprod')
dn = datetime.now()
newname = str(dn.strftime("__%d_%m_%Y__%H_%M__"))
print(newname)
if server=='dev':
    newnamecompleate = 'devoutput'+newname+'.xlsx'
if server == 'prod':
    newnamecompleate = 'prodoutput'+newname+'.xlsx'
#newnamecompleate = 'output'+newname+'.xlsx'
print(newnamecompleate)

# Открываем файл
if server=='dev':
    wb = load_workbook(filename='devautoregress.xlsx', data_only=True)
if server == 'prod':
    wb = load_workbook(filename='prodautoregress.xlsx', data_only=True)
#wb = load_workbook(filename='autoregress.xlsx', data_only=True)
ws = wb.active  # Предполагаем, что данные на активном листе

# Получаем количество строк
N = ws.max_row

# Столбцы от B до I
columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

# Цикл по строкам 2 до N — для каждой строки создаем отдельный payload и отправляем запрос
for i in range(2, N + 1):
    row_dict = {}  # Отдельный dict для текущей строки
    for col in columns:
        # Читаем значения как текст
        val1 = ws[col + '1'].value or ''  # Ключ из header (row 1)
        vali = ws[col + str(i)].value or ''  # Значение из текущей строки

        # Парсим ключ: убираем двоеточие, кавычки и пробелы
        key_str = str(val1).strip().rstrip(':').strip().strip('"\'')  # Пример: '"method": ' -> 'method'
        if not key_str:  # Пропускаем пустые ключи
            continue

        # Значение: оставляем как есть, но можно привести тип
        value_str = str(vali).strip().strip('"\'')  # Убираем внешние кавычки, если они есть

        # Пытаемся определить тип значения (для чисел, булевых)
        try:
            if value_str.lower() in ('true', 'false'):
                value = value_str.lower() == 'true'
            elif value_str.isdigit() or (value_str.startswith('-') and value_str[1:].isdigit()):
                value = int(value_str)
            elif '.' in value_str and value_str.replace('.', '').replace('-', '').isdigit():
                value = float(value_str)
            else:
                # Для строковых значений сначала пробуем парсинг JSON (массивы или объекты)
                if ('{' in value_str or '[' in value_str) and ('}' in value_str or ']' in value_str):
                    try:
                        # Пробуем парсить как JSON напрямую
                        value = json.loads(value_str)
                    except json.JSONDecodeError:
                        # Если не парсится, оставляем строкой
                        value = value_str
                else:
                    value = value_str  # Строка по умолчанию
        except ValueError:
            value = value_str

        # Добавляем в dict для текущей строки
        row_dict[key_str] = value

    # Теперь row_dict — это payload для текущей строки (НЕ update, а отдельно!)
    payload = row_dict

    # Выводим для отладки (теперь каждый раз свой, если данные отличаются)
    #print(f"Payload для строки {i}: {payload}")

    print(payload)

    #print(params)
    if server == 'dev':
        baseurl = 'https://dev2-api.smartanalytics.io/api/insights/'
    if server == 'prod':
        baseurl = 'https://api.smartanalytics.io/api/insights/'
    valiurl = ws['A' + str(i)].value or ''  # Если None, используем пустую строку
    urlstart = baseurl + str(valiurl) + '/start/'   # Соединяем текст без пробела
    print(urlstart)
    report_id = test_get_request(urlstart, payload, token)
    print(report_id)
    if server == 'dev':
        wsbaseurl = "https://dev2-cloud.smartanalytics.io/ru/#!/index/39/reports/14/"
    if server == 'prod':
        wsbaseurl = "https://cloud.smartanalytics.io/ru/#!/index/39/reports/14/"
    ws['K' + str(i)].value = wsbaseurl+str(report_id)
    time.sleep(5)
    if server == 'dev':
        urlsstat = "https://new-dev2-api-stat.smartanalytics.io/api/insight_log/?"
    if server == 'prod':
        urlsstat = "https://new-api-stat.smartanalytics.io/api/insight_log/?"
    errorstring = "error"
    while True:
        if errorstring.lower() in str(report_id).lower():
            print("error")
            status = 'Error'
            break

        status=get_status(str(report_id), token, urlsstat)

        if status=="True":
            break
        if status == "False":
            continue
        if status == 'Error':
            break
        time.sleep(5)
    if status == "True":
        print("puse 5 min")
        time.sleep(300)
        comparestatus = compare(ws['J' + str(i)].value, ws['K' + str(i)].value)
        ws['L' + str(i)].value = "совпадение "+str(comparestatus)+"%)."
        wb.save(newnamecompleate)
    if status == 'Error':
        ws['L' + str(i)].value = "ошибка выполнения"
        wb.save(newnamecompleate)






wb.save(newnamecompleate)