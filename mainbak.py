from openpyxl import load_workbook
from startgen import test_get_request
from getisightstatus import get_status
import time
import os
token = os.getenv('tokendev')
# Открываем файл
wb = load_workbook(filename='autoregress.xlsx', data_only=True)
ws = wb.active

# Получаем количество строк
N = ws.max_row

# Столбцы от B до H
columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

# Цикл N-1 раз (для строк 2 до N)
for i in range(2, N + 1):
    # Формируем строку для текущей итерации
    parts = []
    for col in columns:
        # Читаем значения из ячеек как текст
        val1 = ws[col + '1'].value or ''  # Если None, используем пустую строку
        vali = ws[col + str(i)].value or ''  # Если None, используем пустую строку
        combined = str(val1) + str(vali)  # Соединяем текст без пробела
        parts.append(combined)

    # Выводим объединенную строку (соединенные значения, разделенные запятыми)
    print(', '.join(parts))
    if i==2:
        params = {"method": "Искать проблемные зоны", "top_dimensions_number": 10, "max_depth": 4,
                  "metric": "conversion_rate", "set_name": "Веб-аналитика 📈",
                  "time_case": {"TMPM": "Этот месяц VS предыдущий"}, "how_start": "В реальном времени",
                  "run_tests": False}
        #print(params)
        payload="{"+",".join(parts)+"}"
        print(payload)
        valiurl = ws['A' + str(i)].value or ''  # Если None, используем пустую строку
        urlstart = 'https://dev2-api.smartanalytics.io/api/insights/' +str(valiurl) + '/start/'   # Соединяем текст без пробела
        print(urlstart)
        report_id = test_get_request(urlstart, params, token)
        print(report_id)
        ws['J' + str(i)].value = "https://dev2-cloud.smartanalytics.io/ru/#!/index/39/reports/14/"+str(report_id)
        time.sleep(5)
        while get_status(str(report_id), token)==False:
            time.sleep(5)
        break




wb.save('output.xlsx')